"""
app.py — Smart Pipeline Scraper (Standalone)
Phase 1: Quick (Sitemap) -> Phase 2: Deep (Crawl) -> Phase 3: Loop (Relentless)
Keeps all input data. CSV Column selection. Clean 3-sheet XLSX.
"""

import streamlit as st
import requests, re, io, time, random, pandas as pd
import xml.etree.ElementTree as ET
from urllib.parse import urljoin, urlparse
from collections import deque
from datetime import datetime
from bs4 import BeautifulSoup
from email_validator import validate_email as ev_validate, EmailNotValidError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib

try:
    import dns.resolver as _dns_resolver
    DNS_AVAILABLE = True
except ImportError:
    DNS_AVAILABLE = False

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE)
TIER1 = re.compile(r"^(editor|admin|press|advert|contact)[a-z0-9._%+\-]*@", re.IGNORECASE)
TIER2 = re.compile(r"^(info|sales|hello|office|team|support|help)[a-z0-9._%+\-]*@", re.IGNORECASE)
BLOCKED_TLDS = ['png','jpg','jpeg','webp','gif','svg','ico','mp4','mp3','pdf','zip','exe','dmg']
PLACEHOLDER_DOMAINS = ['example.com','test.com','domain.com','email.com','placeholder.com']
PLACEHOLDER_LOCALS = ['you','user','name','email','test','example','someone','username']
SUPPRESS_PREFIXES = ['noreply','no-reply','donotreply','bounce','unsubscribe','postmaster','webmaster','daemon']
FREE_EMAIL_DOMAINS = ["gmail.com","yahoo.com","hotmail.com","outlook.com","aol.com","icloud.com"]
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/119.0",
]
PRIORITY_KW = [("contact",100),("write-for-us",95),("guest-post",90),("advertise",88),("about",75),("team",70)]

def is_valid_email(e):
    e = e.strip()
    if not e or e.count('@') != 1: return False
    l, d = e.split('@'); lo, do = l.lower(), d.lower()
    if not l or not d or '.' not in d: return False
    if lo in PLACEHOLDER_LOCALS or do in PLACEHOLDER_DOMAINS: return False
    if any(lo == p or lo.startswith(p) for p in SUPPRESS_PREFIXES): return False
    if do.rsplit('.',1)[-1] in BLOCKED_TLDS: return False
    return True

def tier_key(e):
    if TIER1.match(e): return "1"
    if TIER2.match(e): return "2"
    return "3"
def tier_short(e): return {"1":"Tier 1","2":"Tier 2","3":"Tier 3"}[tier_key(e)]
def sort_by_tier(emails): return sorted(emails, key=tier_key)
def pick_best(emails):
    pool = [e for e in emails if is_valid_email(e)]
    if not pool: return None
    for pat in [TIER1, TIER2]:
        h = [e for e in pool if pat.match(e)]
        if h: return h[0]
    return pool[0]

def make_headers(): return {"User-Agent": random.choice(USER_AGENTS), "Accept": "text/html,*/*;q=0.9"}

def fetch_page(url, timeout=10):
    try:
        r = requests.get(url, headers=make_headers(), timeout=timeout, allow_redirects=True)
        if "text" in r.headers.get("Content-Type","") and r.ok: return r.text, r.status_code
        return None, r.status_code
    except: return None, 0

def extract_emails(html):
    soup = BeautifulSoup(html, "html.parser"); raw = set()
    raw.update(EMAIL_REGEX.findall(soup.get_text(" ")))
    raw.update(EMAIL_REGEX.findall(html))
    for a in soup.find_all("a", href=True):
        if a["href"].lower().startswith("mailto:"):
            raw.add(a["href"][7:].split("?")[0].strip())
    return {e for e in raw if is_valid_email(e)}

def get_links(html, base, domain):
    soup = BeautifulSoup(html, "html.parser"); links = []
    for a in soup.find_all("a", href=True):
        full = urljoin(base, a["href"]); p = urlparse(full)
        if p.netloc == domain and p.scheme in ("http","https"):
            links.append(full.split("#")[0].split("?")[0])
    return list(set(links))

def score_url(url, kws):
    path = urlparse(url).path.lower(); best = 0
    for kw, sc in kws:
        if kw in path: best = max(best, sc - path.count("/")*3)
    return best

def get_sitemap_urls(root_url, limit=4):
    urls = []
    for c in [urljoin(root_url, "/sitemap.xml"), urljoin(root_url, "/sitemap_index.xml")]:
        html, _ = fetch_page(c, timeout=8)
        if not html: continue
        try:
            r_ = ET.fromstring(html); ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
            for loc in r_.findall(".//sm:loc", ns):
                u = loc.text.strip()
                if u.endswith(".xml"):
                    sub, _ = fetch_page(u, timeout=8)
                    if sub:
                        try:
                            for sl in ET.fromstring(sub).findall(".//sm:loc", ns): urls.append(sl.text.strip())
                        except: pass
                else: urls.append(u)
        except: pass
        if urls: break
    if urls:
        sc = sorted([(u, score_url(u, PRIORITY_KW)) for u in urls if score_url(u, PRIORITY_KW) > 0], key=lambda x: -x[1])
        return [u for u, _ in sc][:limit], True
    return [], False

# ── SCRAPER PHASES ────────────────────────────────────────────────────────────
def run_quick_scan(root_url, skip_t1):
    logs = []; t0 = time.time(); domain = urlparse(root_url).netloc
    all_e = set(); visited = set()
    p_urls, _ = get_sitemap_urls(root_url, limit=4)
    if not p_urls: p_urls = [root_url]
    for url in p_urls:
        if url in visited: continue
        visited.add(url)
        logs.append((url, "hit", f"Quick: {urlparse(url).path[:40]}"))
        html, status = fetch_page(url, timeout=8)
        if status == 403:
            logs.append((url, "blocked", "Cloudflare/403 in Quick phase"))
            return {"emails": sort_by_tier(all_e), "time": round(time.time()-t0,1), "pages": len(visited), "blocked": True, "logs": logs}
        if html:
            new = extract_emails(html) - all_e
            all_e.update(new)
            for e in sort_by_tier(new): logs.append((url, "email", e))
        if skip_t1 and any(TIER1.match(e) for e in all_e):
            logs.append((url, "skip", "T1 found, stopping early")); break
    return {"emails": sort_by_tier(all_e), "time": round(time.time()-t0,1), "pages": len(visited), "blocked": False, "logs": logs}

def run_deep_scan(root_url, skip_t1, max_pages=30):
    logs = []; t0 = time.time(); domain = urlparse(root_url).netloc
    all_e = set(); visited = set(); queue = deque()
    p_urls, _ = get_sitemap_urls(root_url, limit=999)
    for u in reversed(p_urls): queue.appendleft((u, 0))
    queue.append((root_url, 0))
    pd_ = 0
    while queue and pd_ < max_pages:
        url, depth = queue.popleft()
        if url in visited: continue
        visited.add(url); pd_ += 1
        logs.append((url, "hit", f"Deep [{pd_}/{max_pages}]: {urlparse(url).path[:40]}"))
        html, status = fetch_page(url, timeout=12)
        if status == 403:
            logs.append((url, "blocked", "Cloudflare/403 in Deep phase"))
            return {"emails": sort_by_tier(all_e), "time": round(time.time()-t0,1), "pages": pd_, "blocked": True, "logs": logs}
        if html:
            new = extract_emails(html) - all_e
            all_e.update(new)
            for e in sort_by_tier(new): logs.append((url, "email", e))
            if depth < 2:
                for link in get_links(html, url, domain):
                    if link not in visited: queue.append((link, depth + 1))
        if skip_t1 and any(TIER1.match(e) for e in all_e):
            logs.append((url, "skip", "T1 found, stopping early")); break
        time.sleep(0.3)
    return {"emails": sort_by_tier(all_e), "time": round(time.time()-t0,1), "pages": pd_, "blocked": False, "logs": logs}

def run_loop_scan(root_url, max_loops=2):
    logs = []; t0 = time.time()
    for i in range(max_loops):
        logs.append((root_url, "loop", f"Loop attempt {i+1}/{max_loops} (randomized delay)"))
        time.sleep(random.uniform(2, 5))
        result = run_deep_scan(root_url, skip_t1=True, max_pages=15)
        logs.extend(result["logs"])
        if result["emails"]:
            logs.append((root_url, "ok", f"Found emails on loop {i+1}"))
            result["time"] = round(time.time()-t0, 1)
            return result
        if result["blocked"]:
            logs.append((root_url, "blocked", "Hard block detected, stopping loops"))
            break
    return {"emails": [], "time": round(time.time()-t0,1), "pages": 0, "blocked": False, "logs": logs}

# ── VALIDATION ENGINE (Lightweight for Auto-Validate) ─────────────────────────
def _val_syntax(email):
    try: ev_validate(email); return True
    except EmailNotValidError: return False

def _val_mx(domain):
    try:
        recs = _dns_resolver.resolve(domain, "MX")
        return True, [str(r.exchange) for r in recs]
    except: return False, []

def validate_email_full(email):
    domain = email.split("@")[-1].lower()
    syntax = _val_syntax(email)
    mx_ok, mx_h = _val_mx(domain) if DNS_AVAILABLE else (False, [])
    spf = False; dmarc = False; mbox = False; ca = False
    if DNS_AVAILABLE and mx_ok:
        try:
            for rd in _dns_resolver.resolve(domain, "TXT"):
                if "v=spf1" in str(rd): spf = True
            for rd in _dns_resolver.resolve(f"_dmarc.{domain}", "TXT"):
                if "v=DMARC1" in str(rd): dmarc = True
        except: pass
        try:
            mx = mx_h[0].rstrip(".")
            with smtplib.SMTP(mx, timeout=6) as s:
                s.helo("example.com"); s.mail("test@example.com")
                code, _ = s.rcpt(email)
                mbox = code == 250
                s.mail("test@example.com")
                code2, _ = s.rcpt(f"randomaddress9x7z@{domain}")
                ca = code2 == 250
        except: pass
    free = domain in FREE_EMAIL_DOMAINS
    if not syntax: st, re = "Not Deliverable", "Invalid syntax"
    elif not mx_ok: st, re = "Not Deliverable", "No MX records"
    elif mbox:
        if free: st, re = ("Risky", "Catch-all + free") if ca else ("Deliverable", "Free provider")
        elif ca: st, re = "Risky", "Catch-all enabled"
        elif not spf: st, re = "Risky", "Missing SPF"
        else: st, re = "Deliverable", "—"
    else:
        if ca: st, re = "Risky", "Catch-all, mailbox unknown"
        elif free: st, re = "Deliverable", "Free provider (unverified)"
        else: st, re = "Deliverable", "MX/SPF OK, mailbox unconfirmed"
    return {"status": st, "reason": re, "spf": spf, "dmarc": dmarc, "catch_all": ca, "free": free}

# ── XLSX BUILDER ──────────────────────────────────────────────────────────────
def _mf(h): return PatternFill("solid", fgColor=h)
def _fn(b=False, c="111111", s=10, n="Calibri"): return Font(bold=b, color=c, size=s, name=n)
def _bd(): t = Side(style="thin", color="E5E7EB"); return Border(left=t, right=t, top=t, bottom=t)
def _ct(): return Alignment(horizontal="center", vertical="center")
def _lt(): return Alignment(horizontal="left", vertical="center", wrap_text=False)
HDR = _mf("111111"); RF_N = _mf("F9FAFB")
EF_D = _mf("DCFCE7"); EF_R = _mf("FEF3C7"); EF_B = _mf("FECACA")
SF = {"Deliverable": _mf("16A34A"), "Risky": _mf("D97706"), "Not Deliverable": _mf("DC2626")}
TF1 = _mf("FEF9C3"); TF2 = _mf("EEF2FF"); TF3 = _mf("F1F5F9")

def _hdr(ws, r, c, v, w=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.fill = HDR; cl.font = _fn(True, "FFFFFF"); cl.alignment = _ct(); cl.border = _bd()
    if w: ws.column_dimensions[get_column_letter(c)].width = w

def _cl(ws, r, c, v, fl=None, fn_=None, al=None):
    cl = ws.cell(row=r, column=c, value=v)
    if fl: cl.fill = fl
    if fn_: cl.font = fn_
    if al: cl.alignment = al
    cl.border = _bd()

def build_xlsx(results, original_columns):
    wb = Workbook()
    ws = wb.active; ws.title = "Results"; ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 26
    
    # Write original CSV columns first (if any)
    for ci, col_name in enumerate(original_columns, 1):
        w = min(max(len(str(col_name)) * 2, 15), 40)
        _hdr(ws, 1, ci, col_name, w=w)

    # Write scraper columns appended to the right
    scrape_cols = [("Domain",22),("Best Email",32),("Tier",9),("Status",16),
                 ("Score",8),("Reason",22),("Phase Found",12),
                 ("Pages",8),("Time(s)",9),("Total Emails",12)]
    val_offset = len(original_columns)
    for ci, (n, w) in enumerate(scrape_cols, val_offset + 1):
        _hdr(ws, 1, ci, n, w=w)

    for ri, row in enumerate(results, 2):
        orig_data = row.get("orig_data", {})
        
        # Write original data
        for ci, col_name in enumerate(original_columns, 1):
            val = orig_data.get(col_name, "")
            try:
                if pd.isna(val): val = ""
            except TypeError: pass
            _cl(ws, ri, ci, val, RF_N, _fn(s=9), _lt())

        # Write scraped data
        v = row.get("val"); st_ = v.get("status","") if v else ""
        rf = SF.get(st_, RF_N); em = row.get("best_email","")
        tf = TF1 if "1" in row.get("tier","") else (TF2 if "2" in row.get("tier","") else TF3)
        
        v_idx = val_offset + 1
        _cl(ws, ri, v_idx, row.get("domain",""), RF_N, _fn(True), _lt())
        _cl(ws, ri, v_idx+1, em, EF_D if st_=="Deliverable" else RF_N, _fn(True,n="Courier New",s=9), _lt())
        _cl(ws, ri, v_idx+2, row.get("tier","—"), tf, _fn(), _ct())
        sf_ = SF.get(st_)
        _cl(ws, ri, v_idx+3, st_ or "—", sf_ or RF_N, _fn(True,c="FFFFFF" if sf_ else "111111"), _ct())
        _cl(ws, ri, v_idx+4, row.get("score","—") if row.get("score") is not None else "—", RF_N, _fn(True), _ct())
        _cl(ws, ri, v_idx+5, v.get("reason","—") if v else "—", RF_N, _fn(s=9), _lt())
        _cl(ws, ri, v_idx+6, row.get("phase","—"), RF_N, _fn(), _ct())
        _cl(ws, ri, v_idx+7, row.get("pages",0), RF_N, _fn(), _ct())
        _cl(ws, ri, v_idx+8, row.get("time",""), RF_N, _fn(), _ct())
        _cl(ws, ri, v_idx+9, "; ".join(row.get("all_emails",[])), RF_N, _fn(n="Courier New",s=8,c="666666"), _lt())

    ws2 = wb.create_sheet("Crawl Log"); ws2.freeze_panes = "A2"
    for ci, (n, w) in enumerate([("Domain",22),("Page URL",50),("Action",10),("Detail",50)], 1): _hdr(ws2, 1, ci, n, w)
    r2 = 2
    for row in results:
        dom = row.get("domain","")
        for log in row.get("logs", []):
            url, act, det = log
            _cl(ws2,r2,1,dom,RF_N,_fn(s=9),_lt())
            _cl(ws2,r2,2,url,RF_N,_fn(s=9),_lt())
            _cl(ws2,r2,3,act,RF_N,_fn(s=9),_ct())
            _cl(ws2,r2,4,det,RF_N,_fn(s=9),_lt())
            ws2.row_dimensions[r2].height = 15; r2+=1

    ws3 = wb.create_sheet("Stats")
    ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 15
    nt = len(results)
    ne = sum(1 for r in results if r.get("best_email"))
    nq = sum(1 for r in results if r.get("phase")=="Quick")
    nd = sum(1 for r in results if r.get("phase")=="Deep")
    nl = sum(1 for r in results if r.get("phase")=="Loop")
    nv = sum(1 for r in results if (r.get("val",{}) or {}).get("status")=="Deliverable")
    nr = sum(1 for r in results if (r.get("val",{}) or {}).get("status")=="Risky")
    nb = sum(1 for r in results if (r.get("val",{}) or {}).get("status")=="Not Deliverable")
    stats = [("Total URLs Processed", nt), ("URLs with Emails", ne), ("Found in Quick Phase", nq),
             ("Required Deep Phase", nd), ("Required Loop Phase", nl), ("Validated Deliverable", nv),
             ("Validated Risky", nr), ("Validated Failed", nb)]
    for i,(l,v) in enumerate(stats, 2):
        _cl(ws3,i,1,l,RF_N,_fn()); _cl(ws3,i,2,v,RF_N,_fn(True),_ct())

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT APP
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Smart Pipeline Scraper", page_icon="🔍", layout="wide", initial_sidebar_state="expanded")
ACCENT = "#111111"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
*,html,body,[class*="css"] {{ font-family:'Inter',system-ui,sans-serif !important; }}
#MainMenu,footer,header {{ visibility:hidden; }}
.block-container {{ padding:1.2rem 2rem 4rem !important; max-width:100% !important; background:#f6f5f2 !important; }}
[data-testid="stSidebar"] {{ background:#111 !important; }}
[data-testid="stSidebar"] * {{ color:#ccc !important; }}
[data-testid="stSidebar"] .stDownloadButton > button {{ background:{ACCENT} !important; border:none !important; color:#fff !important; border-radius:8px !important; font-weight:700 !important; width:100% !important; }}
.mh-ph {{ display:flex; align-items:center; gap:12px; padding:14px 20px; background:#fff; border:1px solid #e8e8e4; border-radius:12px; margin-bottom:16px; }}
.mh-pi {{ width:38px; height:38px; border-radius:10px; background:{ACCENT}; display:flex; align-items:center; justify-content:center; font-size:18px; color:#fff; flex-shrink:0; }}
.mh-pt {{ font-size:17px; font-weight:800; color:#111; letter-spacing:-.4px; }}
.mh-ps {{ font-size:11px; color:#aaa; margin-top:1px; }}
.mh-sec {{ font-size:9.5px; font-weight:700; letter-spacing:1.3px; text-transform:uppercase; color:#c0bfbb; display:block; margin-bottom:6px; }}
.stButton > button {{ font-family:'Inter',sans-serif !important; font-weight:600 !important; border-radius:8px !important; font-size:12.5px !important; height:36px !important; transition:all .13s ease !important; }}
.stButton > button[kind="primary"] {{ background:{ACCENT} !important; border:2px solid {ACCENT} !important; color:#fff !important; }}
.stButton > button[kind="primary"]:hover {{ opacity:.88 !important; }}
.stButton > button[kind="secondary"] {{ background:#fff !important; border:1.5px solid #ddd !important; color:#555 !important; }}
.mh-big .stButton > button {{ height:46px !important; font-size:14px !important; font-weight:800 !important; }}
.stDownloadButton > button {{ font-family:'Inter',sans-serif !important; font-weight:600 !important; border-radius:8px !important; height:36px !important; background:{ACCENT} !important; border:none !important; color:#fff !important; }}
[data-testid="stMetric"] {{ background:#fff; border:1px solid #e8e8e4; border-radius:10px; padding:.75rem .9rem !important; }}
[data-testid="stMetricLabel"] p {{ font-size:9.5px !important; font-weight:700 !important; color:#c0bfbb !important; text-transform:uppercase !important; }}
[data-testid="stMetricValue"] {{ font-size:22px !important; font-weight:800 !important; color:#111 !important; }}
.vp {{ height:4px; background:#f0f0ee; border-radius:99px; overflow:hidden; margin:6px 0; }}
.vf {{ height:100%; border-radius:99px; background:{ACCENT}; transition:width .35s; }}
.mh-log {{ background:#18181b; border-radius:8px; padding:10px 12px; font-family:'JetBrains Mono',monospace; font-size:10.5px; line-height:1.8; max-height:250px; overflow-y:auto; margin-top:6px; }}
.mh-log::-webkit-scrollbar {{ width:4px; }}
.mh-log::-webkit-scrollbar-thumb {{ background:#3f3f46; border-radius:2px; }}
.lr {{ color:#fff; font-weight:700; border-top:1px solid #27272a; margin-top:4px; padding-top:4px; }}
.lr:first-child {{ border-top:none; margin-top:0; padding-top:0; }}
.lo {{ color:#4ade80; font-weight:600; }}
.lf {{ color:#f87171; }}
.li {{ color:#3f3f46; }}
.lx {{ color:#22d3ee; font-weight:700; }}
.mh-info {{ background:#f0fdf4; border:1px solid #bbf7d0; border-radius:8px; padding:8px 13px; font-size:12px; color:#15803d; font-weight:600; margin:4px 0; }}
.mh-warn {{ background:#fff1f2; border:1px solid #fecdd3; border-radius:8px; padding:8px 13px; font-size:12px; color:#be123c; font-weight:600; margin:4px 0; }}
.cp {{ background:#fafaf8; border:1px solid #e8e8e4; border-radius:8px; padding:10px 14px; margin:6px 0; font-size:11.5px; }}
.cp-l {{ font-size:9.5px; font-weight:700; color:#999; text-transform:uppercase; letter-spacing:1px; margin-bottom:4px; }}
.cp-v {{ font-family:'JetBrains Mono',monospace; font-size:11px; color:#333; line-height:1.6; }}
</style>""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="font-size:17px;font-weight:800;color:#fff;letter-spacing:-.3px;margin-bottom:4px">Pipeline Scraper</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:10px;color:#555;margin-bottom:16px">Quick -> Deep -> Loop</div>', unsafe_allow_html=True)
    st.divider()
    res = st.session_state.get("sc_results", [])
    if res:
        orig_cols = st.session_state.get("sc_original_cols", [])
        xlsx = build_xlsx(res, orig_cols)
        st.download_button("Export .xlsx", xlsx, f"scrape_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="sc_xlsx", use_container_width=True)
    st.divider()
    st.markdown('<div style="font-size:9px;color:#333;line-height:1.8">Phase 1: Quick (4 pages max)<br>Phase 2: Deep (30 pages max)<br>Phase 3: Loop (Relentless)<br><br>Upload CSV to keep all<br>original columns intact.</div>', unsafe_allow_html=True)

# ── State ──────────────────────────────────────────────────────────────────────
for k, v in {"sc_results":[],"sc_running":False,"sc_idx":0,"sc_log":[],"sc_queue":[],"sc_phase":"Quick","sc_original_cols":[]}.items():
    if k not in st.session_state: st.session_state[k] = v

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""<div class="mh-ph"><div class="mh-pi">🔍</div><div><div class="mh-pt">Smart Pipeline Scraper</div><div class="mh-ps">CSV column selection · automated 3-phase · keeps all data</div></div></div>""", unsafe_allow_html=True)

# ── Input ─────────────────────────────────────────────────────────────────────
c1, c2 = st.columns([4, 1])
df = None
url_col_name = None

with c1:
    raw = st.text_area("u", label_visibility="collapsed", placeholder="https://site1.com\nhttps://site2.com\nsite3.org", height=100, key="url_in")
with c2:
    up = st.file_uploader("CSV", type=["csv"], label_visibility="collapsed", key="sc_up")

urls = []
orig_data_map = {}
original_columns = []

if up:
    try:
        df = pd.read_csv(io.BytesIO(up.read()))
        cols = list(df.columns)
        url_hints = ["url", "website", "site", "domain", "link", "href"]
        det_col = next((c for c in cols if any(h in c.lower() for h in url_hints)), cols[0])
        
        st.markdown(f'<div class="mh-info">Loaded <strong>{len(df)}</strong> rows · <strong>{len(cols)}</strong> columns</div>', unsafe_allow_html=True)
        
        url_col_name = st.selectbox("Select URL Column", cols, index=cols.index(det_col), key="sc_csv_col")
        original_columns = list(df.columns)
        
        for i, row in df.iterrows():
            u = str(row[url_col_name]).strip()
            if pd.notna(row[url_col_name]) and u:
                if not u.startswith("http"): u = "https://" + u
                if u.startswith("http"):
                    urls.append(u)
                    orig_data_map[u] = row.to_dict()
                    
        st.caption(f"Found {len(urls)} valid URLs in '{url_col_name}'")
    except Exception as e:
        st.error(f"Failed to parse CSV: {e}")

if raw:
    for line in raw.splitlines():
        line = line.strip()
        if line:
            if not line.startswith("http"): line = "https://" + line
            urls.append(line)

urls = list(set(urls))

# ── Settings ──────────────────────────────────────────────────────────────────
if urls:
    if not original_columns:
        st.markdown(f'<div class="mh-info">Loaded <strong>{len(urls)}</strong> unique URLs (from text)</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="mh-info">Loaded <strong>{len(urls)}</strong> unique URLs (from CSV) · <strong>{len(original_columns)}</strong> columns will be kept</div>', unsafe_allow_html=True)

    s1, s2, s3, s4 = st.columns(4)
    with s1: skip_t1 = st.toggle("Stop at Tier 1", value=True, key="sc_t1")
    with s2: do_loop = st.toggle("Enable Looping", value=True, key="sc_loop")
    with s3: max_loops = st.slider("Max Loops", 1, 5, 2, key="sc_ml") if do_loop else 2
    with s4: auto_val = st.toggle("Auto-Validate", value=False, key="sc_av")

# ── Controls ───────────────────────────────────────────────────────────────────
vc1, vc2, vc3 = st.columns([3, 1, 1])
with vc1:
    st.markdown('<div class="mh-big">', unsafe_allow_html=True)
    if not st.session_state.sc_running:
        if st.button(f"Start Pipeline ({len(urls)} URLs)", type="primary", use_container_width=True, disabled=not urls, key="sc_go"):
            st.session_state.sc_results = []; st.session_state.sc_idx = 0; st.session_state.sc_log = []
            st.session_state.sc_running = True; st.session_state.sc_phase = "Quick"
            st.session_state.sc_queue = [{"url": u, "orig_data": orig_data_map.get(u, {})} for u in urls]
            st.session_state.sc_original_cols = original_columns
            st.rerun()
    else:
        if st.button("Stop", type="secondary", use_container_width=True, key="sc_stop"):
            st.session_state.sc_running = False; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
with vc2:
    if st.session_state.sc_results:
        if st.button("Clear", type="secondary", use_container_width=True, key="sc_clr"):
            st.session_state.sc_results = []; st.session_state.sc_log = []; st.rerun()
with vc3:
    st.markdown('<div style="font-size:10.5px;color:#aaa;padding-top:12px">~5-30s/url depending on phase</div>', unsafe_allow_html=True)

# ── Progress & Log ────────────────────────────────────────────────────────────
res = st.session_state.sc_results; q = st.session_state.sc_queue; idx = st.session_state.sc_idx; tot = len(q)
if st.session_state.sc_running and tot > 0:
    pct = round(idx/tot*100, 1); cur = q[idx] if idx < tot else ""
    phase = st.session_state.sc_phase
    pcol = {"Quick":"#16a34a","Deep":"#d97706","Loop":"#dc2626"}.get(phase, "#111")
    st.markdown(f'<div style="font-size:12px;font-weight:700;color:#111;margin:6px 0 2px">'
        f'Phase: <span style="color:{pcol};background:{pcol}15;padding:2px 8px;border-radius:4px;font-size:10px">{phase}</span> · {idx}/{tot} — <code>{cur.get("url","")[:40]}</code></div>'
        f'<div class="vp"><div class="vf" style="width:{pct}%;background:{pcol}"></div></div>'
        f'<div style="font-size:20px;font-weight:800;color:{pcol};text-align:right;margin-top:-4px">{pct}%</div>', unsafe_allow_html=True)

ll = st.session_state.sc_log
if ll:
    h = ""
    for source, action, detail in ll[-100:]:
        if action == "row": h += f'<div class="lr">[ {detail} ]</div>'
        elif action == "hit": h += f'<div class="li">  -> {detail}</div>'
        elif action == "email": h += f'<div class="lo">  @ {detail}</div>'
        elif action == "skip": h += f'<div class="lx">  STOP - {detail}</div>'
        elif action == "blocked": h += f'<div class="lf">  BLOCKED - {detail}</div>'
        elif action == "loop": h += f'<div class="lx">  LOOP - {detail}</div>'
        elif action == "ok": h += f'<div class="lo">  OK - {detail}</div>'
        elif action == "fail": h += f'<div class="lf">  FAIL - {detail}</div>'
    st.markdown(f'<div class="mh-log">{h}</div>', unsafe_allow_html=True)

# ── Metrics & Table ──────────────────────────────────────────────────────────
if res:
    ne = sum(1 for r in res if r.get("best_email"))
    nq = sum(1 for r in res if r.get("phase")=="Quick")
    nd = sum(1 for r in res if r.get("phase")=="Deep")
    nl = sum(1 for r in res if r.get("phase")=="Loop")
    nv = sum(1 for r in res if (r.get("val",{}) or {}).get("status")=="Deliverable")
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Emails Found", ne); m2.metric("Quick Hits", nq)
    m3.metric("Deep Hits", nd); m4.metric("Loop Hits", nl)
    m5.metric("Validated OK", nv)

    rows = []
    for r in res:
        v = r.get("val") or {}; s = v.get("status","")
        rows.append({"URL": r["url"],"Domain": r["domain"],"Email": r.get("best_email","—"),
            "Tier": r.get("tier","—"),"Status": s or "—","Score": r.get("score","—"),
            "Phase": r.get("phase","—"),"Pages": r.get("pages",0),"Time": r.get("time","")})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True, height=min(500, 44+max(len(rows),1)*36),
        column_config={"URL":st.column_config.TextColumn("URL",width=200),"Domain":st.column_config.TextColumn("Domain",width=150),
            "Email":st.column_config.TextColumn("Email",width=220),"Tier":st.column_config.TextColumn("Tier",width=60),
            "Status":st.column_config.TextColumn("Status",width=140),"Score":st.column_config.NumberColumn("Score",width=50),
            "Phase":st.column_config.TextColumn("Phase",width=80),"Pages":st.column_config.NumberColumn("Pages",width=50),
            "Time":st.column_config.NumberColumn("Time",width=50)})

# ── Engine ─────────────────────────────────────────────────────────────────────
if st.session_state.sc_running:
    q = st.session_state.sc_queue; idx = st.session_state.sc_idx; tot = len(q)
    if idx >= tot:
        st.session_state.sc_running = False; st.rerun()
    else:
        item = q[idx]; url = item["url"]; domain = urlparse(url).netloc
        orig_data = item.get("orig_data", {})
        
        st.session_state.sc_log.append(("", "row", f"{domain} (Phase: {st.session_state.sc_phase})"))
        
        result_data = None
        phase = st.session_state.sc_phase

        if phase == "Quick":
            res = run_quick_scan(url, skip_t1)
            st.session_state.sc_log.extend(res["logs"])
            if res["emails"] or res["blocked"]:
                result_data = {"phase": "Quick", **res}
            else:
                st.session_state.sc_log.append((url, "fail", "Quick failed, pushing to Deep"))
                st.session_state.sc_phase = "Deep"

        elif phase == "Deep":
            res = run_deep_scan(url, skip_t1, max_pages=30)
            st.session_state.sc_log.extend(res["logs"])
            if res["emails"] or res["blocked"] or not do_loop:
                result_data = {"phase": "Deep", **res}
            else:
                st.session_state.sc_log.append((url, "fail", "Deep failed, pushing to Loop"))
                st.session_state.sc_phase = "Loop"

        elif phase == "Loop":
            res = run_loop_scan(url, max_loops=max_loops)
            st.session_state.sc_log.extend(res["logs"])
            result_data = {"phase": "Loop", **res}

        if result_data:
            best = pick_best(result_data["emails"]) or ""
            row_obj = {
                "url": url, "domain": domain, "best_email": best,
                "tier": tier_short(best) if best else "—",
                "all_emails": result_data["emails"], "phase": result_data["phase"],
                "pages": result_data["pages"], "time": result_data["time"],
                "logs": result_data["logs"], "val": None, "score": None,
                "orig_data": orig_data
            }
            
            if auto_val and best:
                v = validate_email_full(best)
                row_obj["val"] = v
                sc = 100
                if tier_key(best) == "2": sc -= 10
                if tier_key(best) == "3": sc -= 25
                if not v.get("spf"): sc -= 15
                if v.get("catch_all"): sc -= 20
                if v["status"] == "Risky": sc -= 30
                if v["status"] == "Not Deliverable": sc -= 65
                row_obj["score"] = max(0, sc)
                st.session_state.sc_log.append((url, "ok", f"Validated: {v['status']}"))

            st.session_state.sc_results.append(row_obj)
            st.session_state.sc_idx += 1
            st.session_state.sc_phase = "Quick"

        st.rerun()

if not urls and not res:
    st.markdown("""
    <div style="text-align:center;padding:60px 0">
        <div style="font-size:48px;opacity:.08;margin-bottom:16px">🔍</div>
        <div style="font-size:18px;font-weight:800;color:#111;margin-bottom:10px">Smart Pipeline Scraper</div>
        <div style="font-size:12.5px;color:#aaa;line-height:2;max-width:420px;margin:0 auto">
            <strong>Text Area:</strong> Paste URLs (one per line).<br>
            <strong>Upload CSV:</strong> Pick which column contains the URLs.<br>
            All other CSV columns are kept intact in the XLSX export.<br><br>
            <strong style="color:#16a34a">Phase 1 (Quick):</strong> Sitemap top 4 pages. ~5s per site.<br>
            <strong style="color:#d97706">Phase 2 (Deep):</strong> Auto-triggers if Quick fails.<br>
            <strong style="color:#dc2626">Phase 3 (Loop):</strong> Relentless mode if Deep fails.
        </div>
    </div>""", unsafe_allow_html=True)
